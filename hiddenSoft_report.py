import datetime
import sqlite3
from colors import color_theme as color
import openpyxl
from openpyxl.styles import (PatternFill, Border, Side, Alignment, Font, NamedStyle)
import re

class hiddenSoftReport:

    def __init__(self, cursor, ) -> None:
        
        self.cursor = cursor
    
    def generate(self, path_to_save):
        
        _q = self.cursor.execute("""
            SELECT u.username, h.pc_name, h.address, h.user_fullname, h.software, h.case_number, h.created FROM hidden_software h
            JOIN users u ON h.id_user = u.id
            WHERE NOT h.isDeleted
        """).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Активные ПК'

        ws['A1'].value = '№'
        ws['B1'].value = 'Кто добавил'
        ws['C1'].value = 'Имя ПК'
        ws['D1'].value = 'Расположение'
        ws['E1'].value = 'ФИО владельца ПК'
        ws['F1'].value = 'Установленное ПО'
        ws['G1'].value = 'Номер кейса'
        ws['H1'].value = 'Время добавления'

        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 13
        ws.column_dimensions['H'].width = 18

        line = 2
        for i in _q:
            ws['A' + str(line)].value = line - 1
            ws['B' + str(line)].value = i[0]
            ws['C' + str(line)].value = i[1]
            ws['D' + str(line)].value = i[2]
            ws['E' + str(line)].value = i[3]
            ws['F' + str(line)].value = i[4]
            ws['G' + str(line)].value = i[5]
            ws['H' + str(line)].value = i[6]
            line += 1

        _w = self.cursor.execute("""
            SELECT u.username, h.pc_name, h.address, h.user_fullname, h.software, h.case_number, h.created, h.delete_time, h.delete_reason FROM hidden_software h
            JOIN users u ON h.id_user = u.id
            WHERE h.isDeleted
        """).fetchall()

        ws2 = wb.create_sheet('Удаленные ПК')
        
        ws2['A1'].value = '№'
        ws2['B1'].value = 'Кто добавил'
        ws2['C1'].value = 'Имя ПК'
        ws2['D1'].value = 'Расположение'
        ws2['E1'].value = 'ФИО владельца ПК'
        ws2['F1'].value = 'Установленное ПО'
        ws2['G1'].value = 'Номер кейса'
        ws2['H1'].value = 'Время добавления'
        ws2['I1'].value = 'Время удаления'
        ws2['J1'].value = 'Причина удаления'

        ws2.column_dimensions['A'].width = 3
        ws2.column_dimensions['B'].width = 40
        ws2.column_dimensions['C'].width = 8
        ws2.column_dimensions['D'].width = 40
        ws2.column_dimensions['E'].width = 40
        ws2.column_dimensions['F'].width = 20
        ws2.column_dimensions['G'].width = 13
        ws2.column_dimensions['H'].width = 18
        ws2.column_dimensions['I'].width = 18
        ws2.column_dimensions['J'].width = 50

        line = 2
        for i in _w:
            ws2['A' + str(line)].value = line - 1
            ws2['B' + str(line)].value = i[0]
            ws2['C' + str(line)].value = i[1]
            ws2['D' + str(line)].value = i[2]
            ws2['E' + str(line)].value = i[3]
            ws2['F' + str(line)].value = i[4]
            ws2['G' + str(line)].value = i[5]
            ws2['H' + str(line)].value = i[6]
            ws2['I' + str(line)].value = i[7]
            ws2['J' + str(line)].value = i[8]
            line += 1
        
        wb.save(path_to_save)
            

if __name__ == '__main__':

    db = r'\\x3.corp.motiv\support$\it_lunatic\Разработка\Python\workday.sqlite'
    savePath = './test.xlsx'
    connect = sqlite3.connect(db, check_same_thread=False)
    cursor = connect.cursor()

    report = hiddenSoftReport(cursor)
    report.generate(savePath)