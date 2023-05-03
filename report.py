import datetime
import sqlite3
from colors import color_theme as color
import openpyxl
from openpyxl.styles import (PatternFill, Border, Side, Alignment, Font, NamedStyle)
import re

class date_datetime:
    def __init__(self, date, mode: str = 'date'):
        # Определяем шаблон и переделываем в класс datetime
        self.mode = mode
        if self.mode == 'date':
            if re.match('(\d{1,2}).(\d{1,2}).(\d{4})', date) is not None:
                self.date = datetime.datetime(
                    year = int(date.split('.')[2]),
                    month = int(date.split('.')[1]),
                    day = int(date.split('.')[0])
                )
            
            elif re.match('(\d{4})-(\d{1,2})-(\d{1,2})', date) is not None:
                self.date = datetime.datetime(
                    year = int(date.split('-')[0]),
                    month = int(date.split('-')[1]),
                    day = int(date.split('-')[2])
                )

            elif date == 'today':
                self.date = datetime.datetime.now()

            else:
                print('Неверный формат')

        elif self.mode == 'class':
            self.date = date

        elif self.mode == 'time':
            if re.match('(\d{1,2}).(\d{1,2}).(\d{4}) (\d{2}):(\d{2})', date) is not None:
                self.date = datetime.datetime(
                    year = int(date.split(' ')[0].split('.')[2]),
                    month = int(date.split(' ')[0].split('.')[1]),
                    day = int(date.split(' ')[0].split('.')[0]),
                    hour = int(date.split(' ')[1].split(':')[0]),
                    minute = int(date.split(' ')[1].split(':')[1])
                )
            
            elif re.match('(\d{4})-(\d{1,2})-(\d{1,2}) (\d{2}):(\d{2})', date) is not None:
                self.date = datetime.datetime(
                    year = int(date.split(' ')[0].split('-')[0]),
                    month = int(date.split(' ')[0].split('-')[1]),
                    day = int(date.split(' ')[0].split('-')[2]),
                    hour = int(date.split(' ')[1].split(':')[0]),
                    minute = int(date.split(' ')[1].split(':')[1])
                )

        self.main()

    def main(self):
        # Определяем какой день недели
        self.weekday_list = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
        self.weekday_id = datetime.datetime.isoweekday(self.date)
        self.weekday_name = self.weekday_list[ self.weekday_id - 1 ]
    
        # Определяем является ли день выходным
        if self.weekday_id in [6, 7]:
            self.isDayoff = True
        
        else:
            self.isDayoff = False

        # Возвращаем форматы
        month = self.date.month
        if month < 10:
            month = '0' + str(month)

        day = self.date.day
        if day < 10:
            day = '0' + str(day)
        
        self.format_standart = str(day) + '.' + str(month) + '.' + str(self.date.year)
        self.format_db = str(self.date.year) + '-' + str(month) + '-' + str(day)

        if self.mode in ['class', 'time']:
            hours = self.date.hour
            if hours < 10:
                hours = '0' + str(hours)

            minutes = self.date.minute
            if minutes < 10:
                minutes = '0' + str(minutes)
            self.format_standart += ' ' + str(hours) + ':' + str(minutes)
            self.format_db += ' ' + str(hours) + ':' + str(minutes)

    def add_day(self, days: int = 1):
        delta = datetime.timedelta(days = days)
        self.date += delta
        self.main()

def generate(cursor, start_date, end_date, path_to_save):
    colors = color()
    start_date = date_datetime(start_date).format_db
    end_date   = date_datetime(end_date).format_db

    data = {}

    # Сбор ID пользователей, у которых параметр root.inReport = 1 (т.е. необходимо внести в отчет)
    user_inReport = cursor.execute("""
        SELECT users.username FROM users, root 
        WHERE users.id = root.id_user AND root.inReport = 1
        ORDER BY users.username ASC
    """).fetchall()

    for user in user_inReport:
        data[user[0]] = {}

    # Собираем дни, использованные в периоде
    # Далее они будут прописаны в кортеже для каждого сотрудника (если что это {})
    used_days = []
    days_list = cursor.execute("""
        SELECT start_day FROM report 
        WHERE status IS NOT 'deleted' AND start_day >= '{start_date} 00:00' AND start_day <= '{end_date} 23:59'
        ORDER BY start_day ASC
    """.format(
        start_date = start_date,
        end_date = end_date
    )).fetchall()

    for record in days_list:
        # Сначала разделяем полученное значение с YYYY-MM-DD HH:MM до YYYY-MM-DD
        # А после преобразовываем в DD.MM.YYYY
        date = record[0].split(' ')[0]
        date = date.split('-')[2] + '.' + date.split('-')[1] + '.' + date.split('-')[0]
        if date not in used_days:
            used_days.append(date)
    # print('[=] Список использованных дней:')
    # print(' '.join(used_days))

    # Добавляем в массив с использованными датами даты, которые не были получены из БД
    last_date = date_datetime(days_list[-1][0].split(' ')[0])
    _end_date  = date_datetime(end_date).date
    if last_date.date != _end_date:
        while last_date.date < _end_date:
            last_date.add_day(days = 1)
            if not last_date.isDayoff:
                used_days.append(last_date.format_standart)
    # print('\n[=] Список с датами, после добавления неиспользованных дат')
    # print(' '.join(used_days))

    # Сразу каждому пользователю проставляем дни и время в них на 0, чтобы если записи в этот день не было, то потом не было пустого места
    # Время начала, конца для рабочего дня и обеда проставляем сразу, чтобы потом можно было на других страницах посмотреть детально за каждого сотрудника
    for user in data:
        for day in used_days:
            data[user][day] = {
                'start_day': '',
                'start_dinner': '',
                'end_dinner': '',
                'end_day': '',
                'day_len': 0,
                'status': '',
                'isNotClosed': False
            }

    # Собираем все записи сотрудника за прошедшее время
    for user in data:
        # print(start_date, ' ', end_date)
        record_list = cursor.execute("""
            SELECT report.start_day, report.start_dinner, report.end_dinner, report.end_day, report.day_len, report.status FROM report, users
            WHERE users.id = report.id_user AND users.username = '{username}' AND report.start_day >= '{start_date} 00:00' AND report.start_day <= '{end_date} 23:59' AND report.status IS NOT 'deleted'
        """.format(
            username = user,
            start_date = start_date,
            end_date = end_date
        )).fetchall()
        for record in record_list:
            day = record[0].split(' ')[0]
            day = day.split('-')[2] + '.' + day.split('-')[1] + '.' + day.split('-')[0]
            if day == '2023-03-31':
                print(record)

            start_day = record[0].split(' ')[1]
            start_dinner = record[1]
            if start_dinner:
                start_dinner = record[1].split(' ')[1]
            else:
                start_dinner = ''
            
            end_dinner = record[2]
            if end_dinner:
                end_dinner = record[2].split(' ')[1]
            else:
                end_dinner = ''

            end_day = record[3]
            if end_day:
                end_day = end_day.split(' ')[1]
            else:
                end_day = ''

            # Округляем время, чтобы было удобно Косте
            # 'isNotClosed': False
            day_len = record[4]
            if not day_len:
                data[user][day]['isNotClosed'] = True
            
            else:
                minute = day_len % 60
                sum_time = float(day_len // 60)
                if minute in range(0, 7):
                    minute_to_add = 0.00
                elif minute in range(7, 15) or minute in range(15, 22):
                    minute_to_add = 0.25
                elif minute in range(22, 30) or minute in range(30, 37):
                    minute_to_add = 0.50
                elif minute in range(37, 45) or minute in range(45, 52):
                    minute_to_add = 0.75
                elif minute in range(52, 60):
                    minute_to_add = 1.00

                sum_time += minute_to_add
                if sum_time > 8.00:
                    sum_time = 8.00
                # day_len = str(sum_time).split('.')
                # day_len = day_len[0] + ',' + day_len[1]
                day_len = sum_time

                status = record[5]
                if not status:
                    status = ''
                data[user][day]['start_day'] = start_day
                data[user][day]['start_dinner'] = start_dinner
                data[user][day]['end_dinner'] = end_dinner
                data[user][day]['end_day'] = end_day
                data[user][day]['day_len'] = day_len
                data[user][day]['status'] = status

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Общее'
    
    # Переменные для стилизации
    side = Side(border_style = 'thin')
    border = Border(
        top = side,
        right = side,
        bottom = side,
        left = side
    )

    font = Font(
        name = 'Calibri',
        size = 11
    )

    auto_fill = PatternFill(
        fill_type = 'solid',
        fgColor = colors.red[1:]
    )

    alignment = Alignment(
        horizontal = 'center',
        vertical = 'center'
    )

    main_style = NamedStyle(
        name = 'main',
        font = font,
        border = border,
        alignment = alignment
    )
    wb.add_named_style(main_style)
    # Создаем статичные данные
    ws.merge_cells('A1:A2')
    ws['A1'].value = 'Дата'
    ws['A1'].font = font
    ws['A1'].border = Border(
        top = side,
        left = side,
        right = side
    )
    ws['A1'].alignment = alignment
    ws['A2'].font = font
    ws['A2'].border = Border(
        bottom = side,
        left = side,
        right = side
    )

    ws.merge_cells('B1:B2')
    ws['B1'].value = 'День недели'
    ws['B1'].font = font
    ws['B1'].border = Border(
        top = side,
        left = side,
        right = side
    )
    ws['B1'].alignment = alignment
    ws['B2'].font = font
    ws['B2'].border = Border(
        bottom = side,
        left = side,
        right = side
    )
    ws['B2'].alignment = alignment
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 13.3
    line = 2
    week_list = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']

    table_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    word_list = table_list
    full_words = []

    color_list = ['e17055', 'fdcb6e', 'ffeaa7', '55efc4', '74b9ff', '0984e3', 'a29bfe']

    for i in table_list:
        for k in word_list:
            full_words.append(i + '' + k)

    full_words = table_list + full_words

    for day in used_days:
        line += 1
        ws['A' + str(line)].value = day
        ws['A' + str(line)].style = 'main'

        date = datetime.datetime(
            year = int(day.split('.')[2]),
            month = int(day.split('.')[1]),
            day = int(day.split('.')[0])
        )
        week_id = datetime.datetime.isoweekday(date) - 1

        ws['B' + str(line)].value = week_list[week_id]
        ws['B' + str(line)].style = 'main'

        col = 1
        for user in data:
            col += 1
            ws[full_words[col] + '1'].value = user
            ws[full_words[col] + '1'].style = 'main'
            ws[full_words[col] + '2'].value = 'Отработанное время, ч'
            ws[full_words[col] + '2'].style = 'main'
            ws.column_dimensions[full_words[col]].width = 35

            if data[user][day]['isNotClosed']:
                # Проверка, была ли завершена запись
                # Данное условие выполняется, тогда, когда запись не была закрыта
                ws[full_words[col] + '' + str(line)].value = 'Смена не была закрыта'
            else:
                ws[full_words[col] + '' + str(line)].value = data[user][day]['day_len']
            ws[full_words[col] + '' + str(line)].style = 'main'
            ws[full_words[col] + '' + str(line)].number_format = '0.00'
            if data[user][day]['status'] == 'auto':
                ws[full_words[col] + '' + str(line)].fill = auto_fill
    
    line += 1
    ws.merge_cells('A' + str(line) + ':' + 'B' + str(line))
    ws['A' + str(line)].border = Border(
        left = side,
        top = side,
        bottom = side
    )
    ws['B' + str(line)].border = Border(
        top = side,
        right = side,
        bottom = side
    )
    ws['A' + str(line)].value = 'Сумма часов:'
    ws['A' + str(line)].alignment = alignment

    color_id = 0
    for c in range(2, col + 1):
        ws[full_words[c] + '' + str(line)].value = '=SUM(' + full_words[c] + '3:' + full_words[c] + '' + str(line - 1) + ')'
        ws[full_words[c] + '' + str(line)].style = 'main'
        ws[full_words[c] + '' + str(line)].number_format = '0.00'
        ws[full_words[c] + '' + str(line)].fill = PatternFill(
            fill_type = 'solid',
            fgColor = color_list[color_id]
        )
        color_id += 1

    start_date = date_datetime(start_date).format_standart
    end_date = date_datetime(str(end_date).split(' ')[0]).format_standart
    wb.save(path_to_save + '/Отчет за период ' + start_date + ' - ' + end_date + '.xlsx')

if __name__ == '__main__':
    connect = sqlite3.connect(r"\\x3.corp.motiv\support$\it_lunatic\Разработка\Python\workday.sqlite")
    cursor = connect.cursor()
    save_path = r'\\x3\support$\it_lunatic'
    generate(
        cursor = cursor,
        start_date = '2023-03-01',
        end_date = '2023-03-31',
        path_to_save = save_path
    )
    print('[=] Отчет успешно сгенерирован')